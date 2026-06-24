import io
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, jsonify, send_file
from core.db import get_db
from core.utils import load_query, get_param_map, _load_backlog

syndications_bp = Blueprint("syndications", __name__)

@syndications_bp.route("/syndications")
def syndications_list():
    conn = get_db()
    syndications = conn.execute(load_query("syndications_list")).fetchall()
    fec_map = get_param_map("FEC", conn)
    status_map = get_param_map("Status", conn)
    customers = conn.execute("SELECT Customerid, CustomerName FROM BOA.CUS.Customer WHERE IsStructured=1 ORDER BY CustomerName").fetchall()
    conn.close()
    return render_template("syndications/syndications.html", syndications=syndications, fec_map=fec_map, status_map=status_map, customers=customers)

@syndications_bp.route("/syndications/add", methods=["POST"])
def add_syndication():
    conn = get_db()
    cid = int(request.form["customerid"])
    prod_code = "SYNDICATION"
    amt = float(request.form["amount"])
    pricing = float(request.form["pricing"]) if request.form.get("pricing") else None
    fec = int(request.form["fec"]) if request.form.get("fec") else 0
    status = request.form["status"]
    exp_date = request.form["expected_date"] if request.form.get("expected_date") else None
    
    conn.execute("INSERT INTO BOA.STR.MainDeals (ProductCode, CustomerId) VALUES (?, ?)", (prod_code, cid))
    did = conn.execute("SELECT MAX(DealId) AS id FROM BOA.STR.MainDeals").fetchone()["id"]
    conn.execute(
        "INSERT INTO BOA.STR.Syndication (DealId, Amount, Pricing, FEC, Status, ExpectedDate) VALUES (?, ?, ?, ?, ?, ?)",
        (did, amt, pricing, fec, status, exp_date)
    )
    conn.commit()
    conn.close()
    return redirect(url_for("syndications.syndications_list"))

@syndications_bp.route("/syndications/<int:deal_id>")
def syndication_detail(deal_id):
    conn = get_db()
    syn = conn.execute(
        "SELECT m.DealId, m.ProductCode, c.CustomerName, s.Amount, s.Pricing, s.FEC, s.Status, s.ExpectedDate "
        "FROM BOA.STR.MainDeals m "
        "JOIN BOA.STR.Syndication s ON m.DealId = s.DealId "
        "JOIN BOA.CUS.Customer c ON m.CustomerId = c.Customerid "
        "WHERE m.DealId = ?", (deal_id,)
    ).fetchone()

    if not syn:
        conn.close()
        return redirect(url_for("syndications.syndications_list"))

    details      = conn.execute("SELECT * FROM BOA.STR.SyndicationBanks WHERE DealId = ?", (deal_id,)).fetchall()
    fec_map      = get_param_map("FEC", conn)
    status_map   = get_param_map("Status", conn)
    # Work items portal
    wit_items, wit_prereq_map, wit_subitems_map, wit_assignees_map = _load_backlog(conn, "syndication", deal_id)
    stakeholders = conn.execute("SELECT StakeholderID, FullName FROM BOA.COR.Stakeholder WHERE IsActive=1 ORDER BY FullName").fetchall()
    users        = conn.execute("SELECT id, username, surname FROM BOA.COR.[User] ORDER BY username").fetchall()
    conn.close()
    return render_template(
        "syndications/syndication_detail.html",
        syn=syn,
        details=details,
        fec_map=fec_map,
        status_map=status_map,
        wit_items=wit_items,
        wit_prereq_map=wit_prereq_map,
        wit_subitems_map=wit_subitems_map,
        wit_assignees_map=wit_assignees_map,
        stakeholders=stakeholders,
        users=users
    )

@syndications_bp.route("/syndications/<int:deal_id>/detail", methods=["POST"])
def add_syndication_detail(deal_id):
    conn = get_db()
    bank_name = request.form["bank_name"]
    amount = float(request.form["amount"]) if request.form.get("amount") else None
    offer_pricing = float(request.form["offer_pricing"]) if request.form.get("offer_pricing") else None
    
    conn.execute(
        "INSERT INTO BOA.STR.SyndicationBanks (DealId, BankName, Amount, OfferPricing) VALUES (?, ?, ?, ?)",
        (deal_id, bank_name, amount, offer_pricing)
    )
    conn.commit()
    conn.close()
    return redirect(url_for("syndications.syndication_detail", deal_id=deal_id))

@syndications_bp.route("/api/syndications/<int:deal_id>", methods=["PATCH"])
def api_syndication_update(deal_id):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "UPDATE BOA.STR.Syndication SET Amount=?, Pricing=?, FEC=?, Status=?, ExpectedDate=? WHERE DealId=?",
        (data.get("amount"), data.get("pricing"), data.get("fec"), data.get("status"), data.get("expected_date") or None, deal_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@syndications_bp.route("/api/syndications/<int:deal_id>", methods=["DELETE"])
def api_syndication_delete(deal_id):
    conn = get_db()
    conn.execute("DELETE FROM BOA.STR.SyndicationBanks WHERE DealId=?", (deal_id,))
    conn.execute("DELETE FROM BOA.STR.Syndication WHERE DealId=?", (deal_id,))
    conn.execute("DELETE FROM BOA.STR.MainDeals WHERE DealId=?", (deal_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@syndications_bp.route("/api/syndications/<int:deal_id>/banks/<int:bank_id>", methods=["PATCH"])
def api_syndication_bank_update(deal_id, bank_id):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "UPDATE BOA.STR.SyndicationBanks SET BankName=?, Amount=?, OfferPricing=? WHERE DealDetailId=? AND DealId=?",
        (data.get("bank_name"), data.get("amount"), data.get("offer_pricing"), bank_id, deal_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@syndications_bp.route("/api/syndications/<int:deal_id>/banks/<int:bank_id>", methods=["DELETE"])
def api_syndication_bank_delete(deal_id, bank_id):
    conn = get_db()
    conn.execute(
        "DELETE FROM BOA.STR.SyndicationBanks WHERE DealDetailId=? AND DealId=?",
        (bank_id, deal_id)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@syndications_bp.route("/syndications/export")
def export_syndications():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    conn = get_db()
    syndications = conn.execute(load_query("syndications_list")).fetchall()
    fec_map = get_param_map("FEC", conn)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Syndications"

    headers = ["Deal ID", "Company Name", "Product", "Amount", "Pricing", "Currency", "Status", "Expected Date"]
    hfont  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill  = PatternFill(start_color="1D62F1", end_color="1D62F1", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, border

    for ri, d in enumerate(syndications, 2):
        vals = [
            d["DealId"], d["CustomerName"], d["ProductCode"], d["Amount"], d["Pricing"],
            fec_map.get(str(d["FEC"] or 0), {}).get("description", "TRY"),
            d["Status"], d["ExpectedDate"]
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"syndications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
