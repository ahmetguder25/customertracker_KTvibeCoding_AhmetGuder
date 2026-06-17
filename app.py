"""Flask application for Customer Tracking System."""
import os
import platform
import json
from datetime import datetime, timedelta
from flask import (Flask, render_template, request, redirect, url_for,
                   send_file, flash, jsonify, session, has_request_context,
                   Response, stream_with_context)
from werkzeug.utils import secure_filename
import io

# ── App Setup ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = "customer-tracker-secret-key-2026"

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
QUERY_DIR  = os.path.join(BASE_DIR, "queries")

UPLOAD_FOLDER     = os.path.join(BASE_DIR, "static", "logos")
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "svg"}
MAX_LOGO_SIZE      = 2 * 1024 * 1024  # 2 MB

PRODUCT_DOCS_FOLDER = os.path.join(BASE_DIR, "static", "product_docs")
os.makedirs(PRODUCT_DOCS_FOLDER, exist_ok=True)


# ── SQL Query Loader ───────────────────────────────────────────────────────────

def load_query(name: str, query_dir: str = None) -> str:
    """Load SQL from <query_dir>/<name>.sql.  Defaults to the main queries/ dir.
    Pass an explicit query_dir (e.g. admin/queries/) to load module-specific SQL.
    No caching — edit files live."""
    directory = query_dir if query_dir is not None else QUERY_DIR
    path = os.path.join(directory, name + ".sql")
    with open(path, "r", encoding="utf-8") as f:
        return f.read().strip()


# ── Database Abstraction ───────────────────────────────────────────────────────


class _ProdCursorWrapper:
    """Wraps a pyodbc cursor to expose dict-like rows (matching sqlite3.Row)."""

    def __init__(self, cursor):
        self._cur = cursor

    def _to_dict(self, row):
        cols = [d[0] for d in self._cur.description]
        return dict(zip(cols, row))

    def fetchall(self):
        return [self._to_dict(r) for r in self._cur.fetchall()]

    def fetchone(self):
        row = self._cur.fetchone()
        return self._to_dict(row) if row is not None else None

    def __getattr__(self, name):
        return getattr(self._cur, name)


class DbConnection:
    """pyodbc connection wrapper for LOCAL and PROD SQL Server.

    Both environments use pyodbc; rows are returned as dicts via _ProdCursorWrapper.
    """

    def __init__(self, raw_conn):
        self._conn = raw_conn

    def execute(self, sql: str, params=()):
        cur = self._conn.cursor()
        cur.execute(sql, params)
        return _ProdCursorWrapper(cur)

    def commit(self):
        self._conn.commit()

    def close(self):
        self._conn.close()

    # context-manager support
    def __enter__(self):
        return self

    def __exit__(self, *_):
        self.close()


class _PymssqlDbConnection(DbConnection):
    """pymssql-backed connection wrapper.

    pymssql uses %s placeholders; this subclass converts ? → %s transparently
    so all caller code stays identical to the pyodbc path.
    """

    def execute(self, sql: str, params=()):
        sql_converted = sql.replace("?", "%s")
        cur = self._conn.cursor()
        cur.execute(sql_converted, params if params else ())
        return _ProdCursorWrapper(cur)


def get_db() -> DbConnection:
    """Return the local SQL Server connection (Deals, CustomerDetail, cached Customer)."""
    return _get_db_local()


def get_customer_db() -> DbConnection:
    """Return connection to SRVDNZ BOA database if PROD, else local SQL Server."""
    env = session.get("env", "local") if has_request_context() else "local"
    if env == "prod":
        return _get_db_prod()
    return _get_db_local()


def _get_db_local() -> DbConnection:
    """LOCAL connection — autocommit=False."""
    return _make_local_conn(autocommit=False)


def _get_db_prod() -> DbConnection:
    """PROD connection with autocommit=False (for regular reads/writes)."""
    return _make_prod_conn(autocommit=False)


def _get_db_prod_autocommit() -> DbConnection:
    """PROD connection with autocommit=True (required for multi-statement batches with temp tables)."""
    return _make_prod_conn(autocommit=True)


def _make_local_conn(autocommit: bool = False) -> DbConnection:
    """OS-aware factory: connection to the local SQL Server instance.

    - macOS / Linux  → Docker container, pymssql + SQL Auth (sa + config.json)
                       No system ODBC library required — pymssql bundles FreeTDS.
    - Windows        → Installed SQL Server Express/Developer, pyodbc + Windows Auth
    """
    config_path = os.path.join(BASE_DIR, "config.json")
    config: dict = {}
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config = json.load(f)
        except Exception as exc:
            print(f"Warning: Failed to load config.json: {exc}")

    db_name = config.get("LOCAL_DB_NAME", "BOA")
    os_name = platform.system()

    if os_name == "Windows":
        # Windows: pyodbc + Windows Auth (ODBC built into every Windows install)
        try:
            import pyodbc  # noqa: PLC0415
        except ImportError as exc:
            raise RuntimeError("pyodbc is not installed. Run: pip install pyodbc") from exc

        server = config.get("LOCAL_WIN_SERVER", r".\SQLEXPRESS")
        drivers = [d for d in pyodbc.drivers() if "SQL Server" in d]
        if not drivers:
            raise RuntimeError(
                "No SQL Server ODBC driver found. "
                "Install 'ODBC Driver 17/18 for SQL Server' from Microsoft."
            )
        driver = next((d for d in drivers if "18" in d), None) or drivers[0]
        conn_str = (
            f"DRIVER={{{driver}}};"
            f"SERVER={server};"
            f"DATABASE={db_name};"
            "Trusted_Connection=yes;"
        )
        try:
            raw = pyodbc.connect(conn_str, autocommit=autocommit, timeout=10)
            return DbConnection(raw)
        except Exception as exc:
            raise RuntimeError(f"LOCAL connection failed ({server}/{db_name}): {exc}") from exc

    else:
        # macOS / Linux: pymssql — no system ODBC library required
        try:
            import pymssql  # noqa: PLC0415
        except ImportError as exc:
            raise RuntimeError("pymssql is not installed. Run: pip3 install pymssql") from exc

        server_full = config.get("LOCAL_SERVER", "localhost,1433")
        server_addr = server_full.replace(",", ":")  # pymssql uses 'host:port'
        sa_user = config.get("LOCAL_SA_USER", "sa")
        sa_pass = config.get("LOCAL_SA_PASSWORD", "")
        try:
            raw = pymssql.connect(
                server=server_addr,
                user=sa_user,
                password=sa_pass,
                database=db_name,
                login_timeout=10,
            )
            if autocommit:
                raw.autocommit(True)
            return _PymssqlDbConnection(raw)
        except Exception as exc:
            raise RuntimeError(f"LOCAL connection failed ({server_full}/{db_name}): {exc}") from exc


def _make_prod_conn(autocommit: bool = False) -> DbConnection:
    """Internal factory: build a pyodbc connection to SRVDNZ/BOA."""
    try:
        import pyodbc  # noqa: PLC0415 — optional dependency
    except ImportError as exc:
        raise RuntimeError(
            "pyodbc is not installed. Run: pip install pyodbc"
        ) from exc

    server  = "SRVDNZ"
    db_name = ""
    config_path = os.path.join(BASE_DIR, "config.json")
    if os.path.exists(config_path):
        try:
            with open(config_path, "r", encoding="utf-8") as f:
                config = json.load(f)
                server = config.get("PROD_SERVER", server)
                db_name = config.get("PROD_DB_NAME", db_name)
        except Exception as exc:
            print(f"Warning: Failed to load config.json: {exc}")

    # Auto-detect installed SQL Server ODBC driver (prefer 18, then 17)
    drivers = [d for d in pyodbc.drivers() if "SQL Server" in d]
    if not drivers:
        raise RuntimeError(
            "No SQL Server ODBC driver found. "
            "Install 'ODBC Driver 17 for SQL Server' or '18 for SQL Server'."
        )
    driver = next((d for d in drivers if "18" in d), None) or drivers[0]

    conn_str = (
        f"DRIVER={{{driver}}};"
        f"SERVER={server};"
        + (f"DATABASE={db_name};" if db_name else "")
        + "Trusted_Connection=yes;"
    )
    try:
        raw = pyodbc.connect(conn_str, autocommit=autocommit, timeout=10)
        return DbConnection(raw)
    except Exception as exc:
        raise RuntimeError(f"PROD connection failed ({server}/{db_name}): {exc}") from exc


# ── Helpers ────────────────────────────────────────────────────────────────────

def to_tr_time(t_str):
    """Converts UTC SQLite timestamp to Istanbul time (+3) → YYYY-MM-DD HH:MM"""
    if not t_str:
        return ""
    try:
        dt = datetime.strptime(str(t_str)[:19].replace("T", " "), "%Y-%m-%d %H:%M:%S")
        dt += timedelta(hours=3)
        return dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(t_str)[:16]


app.jinja_env.filters["tr_time"] = to_tr_time


def _fmt_dt(v, chars):
    """Format a date/datetime value (datetime obj or string) to a fixed-length string.
    Works with both pymssql datetime objects and plain strings."""
    if not v:
        return ""
    try:
        return str(v)[:chars]
    except Exception:
        return ""

app.jinja_env.filters["fmtdate"]     = lambda v: _fmt_dt(v, 10)   # → YYYY-MM-DD
app.jinja_env.filters["fmtdatetime"] = lambda v: _fmt_dt(v, 16)   # → YYYY-MM-DD HH:MM


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def get_param_map(param_type, conn=None):
    """Load parameter dict keyed by ParamCode for a given ParamType."""
    close_conn = conn is None
    if close_conn:
        conn = get_db()
    lang_id = session.get("lang", 0) if has_request_context() else 0
    rows = conn.execute(load_query("get_parameters"), (param_type, lang_id)).fetchall()
    if close_conn:
        conn.close()
    return {
        str(r["ParamCode"]): {
            "code":        str(r["ParamCode"]),
            "description": r["ParamDescription"],
            "bg":          r["ParamValue"]  or "bg-gray-500/20",
            "text":        r["ParamValue2"] or "text-gray-400",
            "logo":        r["ParamValue3"] or "",
        }
        for r in rows
    }


# ── Context Processors ─────────────────────────────────────────────────────────

@app.context_processor
def inject_lang_dict():
    """Inject lang_dict + current_lang into all templates."""
    if not has_request_context():
        return dict(lang_dict={}, current_lang=0)
    lang_id = session.get("lang", 0)
    try:
        conn = get_db()
        rows = conn.execute(load_query("get_dictionary"), (lang_id,)).fetchall()
        conn.close()
        lang_dict = {row["Id"]: row["Description"] for row in rows}
    except Exception:
        lang_dict = {}
    return dict(lang_dict=lang_dict, current_lang=lang_id)


# ── User Guard ──────────────────────────────────────────────────────────

@app.before_request
def require_user_selection():
    """Redirect to user-login if no user has been selected yet."""
    if not request.endpoint or request.endpoint in ("user_login", "set_user", "static") or request.endpoint.startswith("admin."):
        return
    if "user_id" not in session:
        return redirect(url_for("user_login"))

# ── Environment Guard ──────────────────────────────────────────────────────────

@app.before_request
def require_env_selection():
    """Redirect to env-login if no environment has been selected yet."""
    if not request.endpoint or request.endpoint in ("user_login", "set_user", "env_login", "set_env", "disconnect", "static") or request.endpoint.startswith("admin."):
        return
    if "env" not in session:
        return redirect(url_for("env_login"))

# ── Frame-mode redirect passthrough ───────────────────────────────────────────

@app.after_request
def preserve_frame_on_redirect(response):
    """If the request came from an iframe (_frame=1), keep it on redirects."""
    is_frame = request.args.get('_frame') == '1' or request.form.get('_frame') == '1'
    if is_frame and response.status_code in (301, 302, 303, 307, 308):
        loc = response.headers.get('Location', '')
        if loc and '_frame=1' not in loc:
            sep = '&' if '?' in loc else '?'
            response.headers['Location'] = loc + sep + '_frame=1'
    return response

# ── Authentication Routes ───────────────────────────────────────────────

@app.route("/")
def index():
    if "user_id" not in session:
        return redirect(url_for("user_login"))
    if "env" not in session:
        return redirect(url_for("env_login"))
    return redirect(url_for("dashboard"))


@app.route("/user-login")
def user_login():
    try:
        conn = get_db()
        users = conn.execute("SELECT * FROM BOA.COR.[User] ORDER BY username").fetchall()
        conn.close()
    except RuntimeError as exc:
        flash(str(exc), "error")
        users = []
    return render_template("user_login.html", users=users)


@app.route("/set-user", methods=["POST"])
def set_user():
    user_id = request.form.get("user_id")
    if not user_id:
        return redirect(url_for("user_login"))

    try:
        conn = get_db()
        user = conn.execute("SELECT * FROM BOA.COR.[User] WHERE id = ?", (user_id,)).fetchone()
        conn.close()
    except RuntimeError:
        return redirect(url_for("user_login"))

    if user:
        session["user_id"] = user["id"]
        session["username"] = user["username"]
        try:
            session["surname"] = user["surname"] or ""
            session["lang"] = user["default_language"] if user["default_language"] is not None else 0
            session["theme"] = user["default_theme"] if user.get("default_theme") else "dark"
        except Exception:
            # Fallback if columns don't exist during transition
            session["surname"] = ""
            session["lang"] = 0
            session["theme"] = "dark"
            
        return redirect(url_for("env_login"))
    return redirect(url_for("user_login"))

@app.route("/env-login")
def env_login():
    return render_template("env_login.html")


@app.route("/set-env", methods=["POST"])
def set_env():
    env = request.form.get("env", "local")
    if env not in ("local", "prod"):
        env = "local"

    if env == "prod":
        # Try the connection before committing to PROD to surface errors early
        try:
            conn = _get_db_prod()
            conn.close()
        except RuntimeError as exc:
            flash(str(exc), "error")
            return redirect(url_for("env_login"))
    else:
        # Validate LOCAL connection on selection too
        try:
            conn = _get_db_local()
            conn.close()
        except RuntimeError as exc:
            flash(str(exc), "error")
            return redirect(url_for("env_login"))

    session["env"] = env
    return redirect(url_for("dashboard"))


@app.route("/disconnect")
def disconnect():
    session.pop("env", None)
    session.pop("user_id", None)
    session.pop("username", None)
    session.pop("surname", None)
    session.pop("lang", None)
    return redirect(url_for("user_login"))


# ── Language ───────────────────────────────────────────────────────────────────

@app.route("/set_language/<int:lang_id>")
def set_language(lang_id):
    if lang_id in (0, 1):
        session["lang"] = lang_id
        user_id = session.get("user_id")
        if user_id:
            conn = get_db()
            conn.execute("UPDATE BOA.COR.[User] SET default_language = ? WHERE id = ?", (lang_id, user_id))
            conn.commit()
            conn.close()
    return redirect(request.referrer or url_for("dashboard"))


# ── Theme ──────────────────────────────────────────────────────────────────────

@app.route("/set_theme/<theme>")
def set_theme(theme):
    if theme in ("dark", "light"):
        session["theme"] = theme
        user_id = session.get("user_id")
        if user_id:
            try:
                conn = get_db()
                conn.execute("UPDATE BOA.COR.[User] SET default_theme = ? WHERE id = ?", (theme, user_id))
                conn.commit()
                conn.close()
            except Exception:
                pass # Ignore if column doesn't exist yet
    return redirect(request.referrer or url_for("dashboard"))




# ── Main App ───────────────────────────────────────────────────────────────────

@app.route("/app")
def dashboard():
    return render_template(
        "dashboard.html",
        active_env   = session.get("env", "local"),
    )

@app.route("/dashboard")
def dashboard_redirect():
    return redirect(url_for("dashboard"))



# ── Syndications ───────────────────────────────────────────────────────────────

@app.route("/syndications")
def syndications_list():
    conn = get_db()
    syndications = conn.execute(load_query("syndications_list")).fetchall()
    fec_map = get_param_map("FEC", conn)
    status_map = get_param_map("Status", conn)
    customers = conn.execute("SELECT Customerid, CustomerName FROM BOA.CUS.Customer WHERE IsStructured=1 ORDER BY CustomerName").fetchall()
    conn.close()
    return render_template("syndications.html", syndications=syndications, fec_map=fec_map, status_map=status_map, customers=customers)

@app.route("/syndications/add", methods=["POST"])
def add_syndication():
    conn = get_db()
    cid = int(request.form["customerid"])
    prod_code = "SYNDICATION"
    amt = float(request.form["amount"])
    pricing = float(request.form["pricing"]) if request.form.get("pricing") else None
    fec = int(request.form["fec"]) if request.form.get("fec") else 0
    status = request.form["status"]
    exp_date = request.form["expected_date"] if request.form.get("expected_date") else None
    
    conn.execute("INSERT INTO BOA.STR.MainDeals (ProductCode, CustomerId) VALUES (?, ?)", (prod_code, cid))
    did = conn.execute("SELECT MAX(DealId) AS id FROM BOA.STR.MainDeals").fetchone()["id"]
    conn.execute(
        "INSERT INTO BOA.STR.Syndication (DealId, Amount, Pricing, FEC, Status, ExpectedDate) VALUES (?, ?, ?, ?, ?, ?)",
        (did, amt, pricing, fec, status, exp_date)
    )
    conn.commit()
    conn.close()
    return redirect(url_for("syndications_list"))

@app.route("/syndications/<int:deal_id>")
def syndication_detail(deal_id):
    conn = get_db()
    syn = conn.execute(
        "SELECT m.DealId, m.ProductCode, c.CustomerName, s.Amount, s.Pricing, s.FEC, s.Status, s.ExpectedDate "
        "FROM BOA.STR.MainDeals m "
        "JOIN BOA.STR.Syndication s ON m.DealId = s.DealId "
        "JOIN BOA.CUS.Customer c ON m.CustomerId = c.Customerid "
        "WHERE m.DealId = ?", (deal_id,)
    ).fetchone()
    
    if not syn:
        conn.close()
        return redirect(url_for("syndications_list"))
        
    details = conn.execute("SELECT * FROM BOA.STR.SyndicationBanks WHERE DealId = ?", (deal_id,)).fetchall()
    fec_map = get_param_map("FEC", conn)
    status_map = get_param_map("Status", conn)
    conn.close()
    return render_template("syndication_detail.html", syn=syn, details=details, fec_map=fec_map, status_map=status_map)

@app.route("/syndications/<int:deal_id>/detail", methods=["POST"])
def add_syndication_detail(deal_id):
    conn = get_db()
    bank_name = request.form["bank_name"]
    amount = float(request.form["amount"]) if request.form.get("amount") else None
    offer_pricing = float(request.form["offer_pricing"]) if request.form.get("offer_pricing") else None
    
    conn.execute(
        "INSERT INTO BOA.STR.SyndicationBanks (DealId, BankName, Amount, OfferPricing) VALUES (?, ?, ?, ?)",
        (deal_id, bank_name, amount, offer_pricing)
    )
    conn.commit()
    conn.close()
    return redirect(url_for("syndication_detail", deal_id=deal_id))


@app.route("/api/syndications/<int:deal_id>", methods=["PATCH"])
def api_syndication_update(deal_id):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "UPDATE BOA.STR.Syndication SET Amount=?, Pricing=?, FEC=?, Status=?, ExpectedDate=? WHERE DealId=?",
        (data.get("amount"), data.get("pricing"), data.get("fec"), data.get("status"), data.get("expected_date") or None, deal_id)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/syndications/<int:deal_id>", methods=["DELETE"])
def api_syndication_delete(deal_id):
    conn = get_db()
    # Delete child banks first, then the syndication, then the main deal
    conn.execute("DELETE FROM BOA.STR.SyndicationBanks WHERE DealId=?", (deal_id,))
    conn.execute("DELETE FROM BOA.STR.Syndication WHERE DealId=?", (deal_id,))
    conn.execute("DELETE FROM BOA.STR.MainDeals WHERE DealId=?", (deal_id,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/syndications/<int:deal_id>/banks/<int:bank_id>", methods=["PATCH"])
def api_syndication_bank_update(deal_id, bank_id):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "UPDATE BOA.STR.SyndicationBanks SET BankName=?, Amount=?, OfferPricing=? WHERE DealDetailId=? AND DealId=?",
        (data.get("bank_name"), data.get("amount"), data.get("offer_pricing"), bank_id, deal_id)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/syndications/<int:deal_id>/banks/<int:bank_id>", methods=["DELETE"])
def api_syndication_bank_delete(deal_id, bank_id):
    conn = get_db()
    conn.execute(
        "DELETE FROM BOA.STR.SyndicationBanks WHERE DealDetailId=? AND DealId=?",
        (bank_id, deal_id)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/syndications/export")
def export_syndications():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    conn = get_db()
    syndications = conn.execute(load_query("syndications_list")).fetchall()
    fec_map = get_param_map("FEC", conn)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Syndications"

    headers = ["Deal ID", "Company Name", "Product", "Amount", "Pricing", "Currency", "Status", "Expected Date"]
    hfont  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill  = PatternFill(start_color="1D62F1", end_color="1D62F1", fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment, c.border = hfont, hfill, halign, border

    for ri, d in enumerate(syndications, 2):
        vals = [
            d["DealId"], d["CustomerName"], d["ProductCode"], d["Amount"], d["Pricing"],
            fec_map.get(str(d["FEC"] or 0), {}).get("description", "TRY"),
            d["Status"], d["ExpectedDate"]
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf, as_attachment=True,
        download_name=f"syndications_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── Management ─────────────────────────────────────────────────────────────────

@app.route("/management")
def management():
    conn         = get_db()
    sector_map   = get_param_map("Sector", conn)
    customers    = conn.execute(load_query("mgmt_list_customers")).fetchall()
    stakeholders = conn.execute(
        "SELECT * FROM BOA.ZZZ.Stakeholder WHERE IsActive=1 ORDER BY FullName"
    ).fetchall()
    conn.close()
    return render_template("management.html", customers=customers, sector_map=sector_map,
                           stakeholders=stakeholders)


# ── Stakeholder CRUD ───────────────────────────────────────────────────────────

@app.route("/api/stakeholders", methods=["POST"])
def api_stakeholder_create():
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "INSERT INTO BOA.ZZZ.Stakeholder (FullName,Organization,Department,Email) VALUES (?,?,?,?)",
        (data.get("full_name", ""), data.get("organization", ""),
         data.get("department", ""), data.get("email", ""))
    )
    conn.commit()
    sid = conn.execute("SELECT MAX(StakeholderID) AS id FROM BOA.ZZZ.Stakeholder").fetchone()["id"]
    conn.close()
    return jsonify({"ok": True, "stakeholder_id": sid})


@app.route("/api/stakeholders/<int:sid>", methods=["PATCH"])
def api_stakeholder_update(sid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "UPDATE BOA.ZZZ.Stakeholder SET FullName=?,Organization=?,Department=?,Email=? WHERE StakeholderID=?",
        (data.get("full_name"), data.get("organization"), data.get("department"), data.get("email"), sid)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/stakeholders/<int:sid>", methods=["DELETE"])
def api_stakeholder_delete(sid):
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.Stakeholder SET IsActive=0 WHERE StakeholderID=?", (sid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/customer/lookup/<int:account_number>")
def lookup_customer(account_number):
    """Fetch a customer from BOADWH.CUS.Customer on SRVDNZ and check local tracking status.
    Always uses the PROD SQL Server connection (the Fetch button is disabled in TEST via HTML).
    Returns distinct JSON flags so the frontend can show actionable error messages.
    """
    try:
        conn_real = _get_db_prod()  # always hit PROD regardless of session env
        customer = conn_real.execute(load_query("prod_lookup_customer"), (account_number,)).fetchone()
        conn_real.close()

        if customer is None:
            return jsonify({"found": False})

        # Check if already tracked in local SQLite
        conn_local = get_db()
        already = conn_local.execute(
            "SELECT IsStructured FROM BOA.CUS.Customer WHERE Customerid = ?", (account_number,)
        ).fetchone()
        conn_local.close()

        return jsonify({
            "found": True,
            "Customerid": customer.get("Customerid", account_number),
            "CustomerName": customer.get("CustomerName", ""),
            "sector": customer.get("sector", "") or "",
            "branch": customer.get("branch", "") or "",
            "region": customer.get("region", "") or "",
            "value_segment": customer.get("value_segment", "") or "",
            "portfolio_manager": customer.get("portfolio_manager", "") or "",
            "customer_class": customer.get("CustomerClassName", "") or "",
            "IsStructured": 1 if already and already["IsStructured"] else 0
        })

    except RuntimeError as e:
        # Connection problem: pyodbc missing, no ODBC driver, SRVDNZ unreachable
        msg = str(e)
        print(f"[lookup_customer] connection error: {msg}")
        return jsonify({"found": False, "connection_error": True, "error": msg})
    except Exception as e:
        # SQL execution error (permissions, bad query, etc.)
        msg = str(e)
        print(f"[lookup_customer] query error: {msg}")
        return jsonify({"found": False, "query_error": True, "error": msg})


@app.route("/management/customer/add", methods=["POST"])
def add_customer():
    customer_id = int(request.form["Customerid"])
    env = session.get("env", "local") if has_request_context() else "local"
    
    if env == "prod":
        conn_real = get_customer_db()
        customer = conn_real.execute(load_query("real_customer_sync"), (customer_id,)).fetchone()
        conn_real.close()
        
        if customer:
            conn_local = get_db()
            conn_local.execute(load_query("customer_upsert"), (
                customer.get("Customerid", customer_id),
                customer.get("CustomerName", ""),
                customer.get("TotalLimit", 0),
                customer.get("credit_limit_currency", "TRY"),
                customer.get("foreign_trade_volume", 1),
                customer.get("memzuc_151_volume", 1),
                customer.get("memzuc_152_volume", 1),
                customer.get("value_segment", ""),
                customer.get("branch", ""),
                customer.get("region", ""),
                customer.get("portfolio_manager", ""),
                customer.get("CustomerClassName", "")
            ))
            conn_local.commit()
            conn_local.close()
    else:
        conn = get_db()
        conn.execute(load_query("mgmt_set_structured"), (customer_id,))
        conn.commit()
        conn.close()
        
    return redirect(url_for("management"))


@app.route("/management/api/sync/queue", methods=["GET"])
def sync_queue():
    # Available in both TEST (demo) and PROD
    conn_local = get_db()
    customers = conn_local.execute("SELECT Customerid, CustomerName FROM BOA.CUS.Customer WHERE IsStructured = 1").fetchall()
    conn_local.close()
    
    return jsonify({
        "success": True,
        "customers": [{"id": c["Customerid"], "name": c["CustomerName"]} for c in customers]
    })

@app.route("/management/api/sync/batch", methods=["POST"])
def sync_batch():
    env = session.get("env", "local") if has_request_context() else "local"
    data = request.get_json()
    cids = data.get("customer_ids", [])

    if not cids:
        return jsonify({"success": False, "error": "No customers to sync."})

    # TEST environment — clearly blocked, no fake data
    if env != "prod":
        return jsonify({
            "success": False,
            "env_error": True,
            "error": "Update is only available in PROD environment. Switch to PROD to sync real customer data."
        })

    # PROD: connect to SRVDNZ > BOA and run batch sync against BOADWH tables
    try:
        query_text = load_query("real_batch_customer_sync")
        # Count how many times {ids} appears BEFORE replacing (each becomes N ?-marks)
        ids_occurrences = query_text.count("{ids}")
        placeholders = ",".join("?" for _ in cids)
        query_text = query_text.replace("{ids}", placeholders)

        # The batch query uses multi-statement (SET NOCOUNT ON + temp tables).
        # It needs autocommit=True so temp table drops/creates don't conflict with a transaction.
        conn_real = _get_db_prod_autocommit()
        # Pass cids once per {ids} occurrence so param count equals marker count
        params = tuple(cids) * ids_occurrences
        results = conn_real.execute(query_text, params).fetchall()
        conn_real.close()
    except RuntimeError as e:
        return jsonify({"success": False, "connection_error": True, "error": str(e)})
    except Exception as e:
        return jsonify({"success": False, "query_error": True, "error": str(e)})

    if results:
        conn_local = get_db()
        for customer in results:
            conn_local.execute("""
                UPDATE Customer SET
                    CustomerName = ?,
                    credit_limit = ?,
                    credit_limit_currency = ?,
                    foreign_trade_volume = ?,
                    memzuc_151_volume = ?,
                    memzuc_152_volume = ?,
                    value_segment = ?,
                    branch = ?,
                    region = ?,
                    portfolio_manager = ?,
                    CustomerClassName = ?,
                    sector = ?
                WHERE Customerid = ?
            """, (
                customer.get("CustomerName", ""),
                customer.get("TotalLimit", 0),
                customer.get("credit_limit_currency", "TRY"),
                customer.get("foreign_trade_volume", 1),
                customer.get("memzuc_151_volume", 1),
                customer.get("memzuc_152_volume", 1),
                customer.get("value_segment", ""),
                customer.get("branch", ""),
                customer.get("region", ""),
                customer.get("portfolio_manager", ""),
                customer.get("CustomerClassName", ""),
                customer.get("sector", ""),
                customer.get("Customerid", customer.get("CustomerID", 0)),
            ))
        conn_local.commit()
        conn_local.close()
        return jsonify({"success": True, "count": len(results)})

    return jsonify({"success": False, "error": "No matching customers found on SRVDNZ (BOADWH). Verify account numbers are correct."})


@app.route("/management/edit/<int:customer_id>", methods=["GET", "POST"])
def edit_customer(customer_id):
    conn = get_db()
    if request.method == "POST":
        logo_filename = None
        if "logo" in request.files:
            file = request.files["logo"]
            if file and file.filename:
                if not allowed_file(file.filename):
                    flash("Invalid file type.", "error")
                    conn.close()
                    return redirect(url_for("edit_customer", customer_id=customer_id))
                file.seek(0, os.SEEK_END)
                if file.tell() > MAX_LOGO_SIZE:
                    flash("Logo file too large.", "error")
                    conn.close()
                    return redirect(url_for("edit_customer", customer_id=customer_id))
                file.seek(0)
                ext       = file.filename.rsplit(".", 1)[1].lower()
                safe_name = secure_filename(request.form["CustomerName"].lower().replace(" ", "_")) + "." + ext
                file.save(os.path.join(UPLOAD_FOLDER, safe_name))
                logo_filename = safe_name

        if logo_filename:
            conn.execute(load_query("mgmt_update_customer_logo"), (
                request.form["CustomerName"],
                request.form.get("sector", ""),
                request.form.get("portfolio_manager", ""),
                logo_filename,
                customer_id,
            ))
        else:
            conn.execute(load_query("mgmt_update_customer"), (
                request.form["CustomerName"],
                request.form.get("sector", ""),
                request.form.get("portfolio_manager", ""),
                customer_id,
            ))
        conn.commit()
        conn.close()
        return redirect(url_for("edit_customer", customer_id=customer_id))

    customer = conn.execute(load_query("mgmt_get_customer"), (customer_id,)).fetchone()
    if not customer:
        conn.close()
        return redirect(url_for("management"))
    sector_map = get_param_map("Sector", conn)
    conn.close()
    return render_template("edit.html", customer=customer, sector_map=sector_map)


@app.route("/management/customer/delete/<int:customer_id>", methods=["POST"])
def delete_customer(customer_id):
    conn = get_db()
    conn.execute(load_query("mgmt_remove_structured"), (customer_id,))
    conn.commit()
    conn.close()
    return redirect(url_for("management"))


# ── Overview ───────────────────────────────────────────────────────────────────

@app.route("/overview")
def overview():
    conn       = get_db()
    customers  = conn.execute(load_query("overview_list")).fetchall()
    sector_map = get_param_map("Sector", conn)
    conn.close()
    return render_template("overview_list.html", customers=customers, sector_map=sector_map)


@app.route("/overview/<int:customer_id>")
def overview_detail(customer_id):
    conn = get_db()
    customer = conn.execute(load_query("mgmt_get_customer"), (customer_id,)).fetchone()
    if not customer:
        conn.close()
        return redirect(url_for("overview"))

    deals        = conn.execute(load_query("overview_deals"),        (customer_id,)).fetchall()
    comments     = conn.execute(load_query("overview_comments"),     (customer_id,)).fetchall()
    same_sector  = conn.execute(load_query("overview_same_sector"),  (customer["sector"],)).fetchone()["cnt"]
    total_cust   = conn.execute(load_query("overview_total_customers")).fetchone()["cnt"]

    lang_id      = session.get("lang", 0)

    status_map    = get_param_map("Status",   conn)
    deal_type_map = get_param_map("DealType", conn)
    fec_map       = get_param_map("FEC",      conn)
    sector_map    = get_param_map("Sector",   conn)
    
    # Financial items processing
    fin_defs = conn.execute("SELECT * FROM BOA.LNS.FinancialItemDefinition").fetchall()
    allotments = conn.execute("SELECT * FROM BOA.LNS.AllotmentFinancialItems WHERE AllotmentMainId = ? ORDER BY PeriodId", (customer_id,)).fetchall()
    conn.close()

    total_deal_size = sum((d["deal_size"] or 0) for d in deals)

    # Build Financial Tree
    fin_map = {}
    for row in fin_defs:
        d = dict(row)
        # ParentId may be stored as the string 'NULL' instead of Python None — normalize it
        pid = d["ParentId"]
        if pid == 'NULL' or pid == '' or pid == 0:
            d["ParentId"] = None
        else:
            try:
                d["ParentId"] = int(pid)
            except (TypeError, ValueError):
                d["ParentId"] = None
        fin_map[d["FinancialItemDefinitionId"]] = d

    periods = sorted(list(set([a["PeriodId"] for a in allotments])))
    
    for f in fin_map.values():
        f["children"] = []
        f["amounts"] = {p: 0 for p in periods}
        
    for f in fin_map.values():
        pid = f["ParentId"]
        if pid and pid in fin_map:
            fin_map[pid]["children"].append(f)
            
    for a in allotments:
        fid = a["FinancialItemDefinitionId"]
        if fid in fin_map:
            fin_map[fid]["amounts"][a["PeriodId"]] += (a["OriginalValue"] or 0)

    def compute_values(node):
        for child in node["children"]:
            compute_values(child)
            for p in periods:
                node["amounts"][p] += child["amounts"][p]

    roots = [n for n in fin_map.values() if not n["ParentId"]]
    for r in roots:
        compute_values(r)
        
    ordered_fin_nodes = []
    def traverse(node, depth):
        # Determine total value to hide completely empty zeroed branches if necessary, but keep all for accuracy
        node_copy = {
            "id": node["FinancialItemDefinitionId"],
            "code": node["Code"],
            "parent_id": node["ParentId"],
            "name": node["NameInEnglish"] or node["Name"],
            "depth": depth,
            "amounts": node["amounts"],
            "has_children": len(node["children"]) > 0
        }
        ordered_fin_nodes.append(node_copy)
        # Sort children safely by putting strings into comparable format, simple string comparison works for codes
        for child in sorted(node["children"], key=lambda x: str(x["Code"])):
            traverse(child, depth + 1)
            
    for r in sorted(roots, key=lambda x: str(x["Code"])):
        traverse(r, 0)
        
    # Prepare Chart Data for Total Assets
    # In Turkish CoA, Assets are 1 (Current) and 2 (Non-Current) or sum thereof
    chart_periods = [f"Period {p}" for p in periods]
    asset_data = []
    # Try to find nodes '1' and '2' explicitly and sum them, or fallback to simple sum of all active leaves.
    node_1 = next((n for n in fin_map.values() if str(n["Code"]) == "1"), None)
    node_2 = next((n for n in fin_map.values() if str(n["Code"]) == "2"), None)
    
    for p in periods:
        val = 0
        if node_1: val += node_1["amounts"].get(p, 0)
        if node_2: val += node_2["amounts"].get(p, 0)
        if not node_1 and not node_2:
            val = sum(n["amounts"][p] for n in fin_map.values() if n["IsLeaf"] and (str(n["Code"]).startswith("1") or str(n["Code"]).startswith("2")))
        asset_data.append(val)

    return render_template(
        "overview_detail.html",
        customer        = customer,
        deals           = deals,
        comments        = comments,
        status_map      = status_map,
        deal_type_map   = deal_type_map,
        fec_map         = fec_map,
        sector_map      = sector_map,
        total_customers = total_cust,
        same_sector     = same_sector,
        total_deal_size = total_deal_size,
        fin_nodes       = ordered_fin_nodes,
        periods         = periods,
        chart_periods   = chart_periods,
        chart_asset_data= asset_data
    )






# ── Comments ───────────────────────────────────────────────────────────────────

@app.route("/overview/<int:customer_id>/comment", methods=["POST"])
def add_comment(customer_id):
    author  = session.get("username", "System")
    content = request.form.get("content", "").strip()
    if author and content:
        conn = get_db()
        conn.execute(load_query("comment_insert"), (customer_id, author, content))
        conn.commit()
        conn.close()
    return redirect(url_for("overview_detail", customer_id=customer_id))


# ── Blueprint Registration ───────────────────────────────────────────────────────────
from admin.admin_bp import admin_bp  # noqa: E402
app.register_blueprint(admin_bp)

# ── Entry Point ────────────────────────────────────────────────────────────────

def _col_exists(conn, table: str, column: str) -> bool:
    row = conn.execute(
        "SELECT COUNT(*) AS cnt FROM INFORMATION_SCHEMA.COLUMNS "
        "WHERE TABLE_SCHEMA='ZZZ' AND TABLE_NAME=? AND COLUMN_NAME=?",
        (table, column)
    ).fetchone()
    return row and int(row["cnt"]) > 0


def _table_exists(conn, table: str) -> bool:
    row = conn.execute(
        "SELECT COUNT(*) AS cnt FROM INFORMATION_SCHEMA.TABLES "
        "WHERE TABLE_SCHEMA='ZZZ' AND TABLE_NAME=?",
        (table,)
    ).fetchone()
    return row and int(row["cnt"]) > 0


def _ensure_isactive_columns():
    """Idempotently add IsActive to CustomerDeals and Comment."""
    try:
        conn = get_db()
        for table in ("Comment",):
            if not _col_exists(conn, table, "IsActive"):
                conn.execute(f"ALTER TABLE BOA.ZZZ.{table} ADD IsActive TINYINT NOT NULL DEFAULT 1")
                conn.commit()
                print(f"[startup] IsActive column added to BOA.ZZZ.{table}")
            else:
                print(f"[startup] IsActive already present on BOA.ZZZ.{table} — skipping")
        conn.close()
    except Exception as e:
        print(f"[startup] IsActive migration warning (non-fatal): {e}")


def _run_platform_migrations():
    """
    Idempotently create all new platform tables (Products, OKRs, Projects, Backlog).
    Safe to call on every startup.
    """
    conn = get_db()
    try:
        # 510xxxx — Stakeholders
        if not _table_exists(conn, "Stakeholder"):
            conn.execute("""
                CREATE TABLE BOA.ZZZ.Stakeholder (
                    StakeholderID INT IDENTITY(5100001,1) PRIMARY KEY,
                    FullName      NVARCHAR(200) NOT NULL,
                    Organization  NVARCHAR(200),
                    Department    NVARCHAR(200),
                    Email         NVARCHAR(200),
                    IsActive      TINYINT NOT NULL DEFAULT 1,
                    CreatedAt     DATETIME NOT NULL DEFAULT GETDATE()
                )""")
            conn.commit()
            print("[startup] Created BOA.ZZZ.Stakeholder")
        else:
            print("[startup] Stakeholder exists — skipping")

        # 110xxxx — Product
        if not _table_exists(conn, "Product"):
            conn.execute("""
                CREATE TABLE BOA.COR.Product (
                    ProductID           INT IDENTITY(1,1) PRIMARY KEY,
                    ProductCode         NVARCHAR(50)  NOT NULL,
                    ProductName         NVARCHAR(200) NOT NULL,
                    IsActive            TINYINT NOT NULL DEFAULT 1
                )""")
            conn.execute(
                "INSERT INTO BOA.COR.Product (ProductCode,ProductName) VALUES ('SYNDICATION','Syndication')"
            )
            conn.commit()
            print("[startup] Created BOA.COR.Product + seeded Unclassified (ID 1100001)")
        else:
            print("[startup] Product exists — skipping")
        # 310xxxx — Objective
        if not _table_exists(conn, "Objective"):
            conn.execute("""
                CREATE TABLE BOA.ZZZ.Objective (
                    ObjectiveID INT IDENTITY(3100001,1) PRIMARY KEY,
                    Title       NVARCHAR(300) NOT NULL,
                    Description NVARCHAR(MAX),
                    Period      NVARCHAR(50),
                    Owner       NVARCHAR(100),
                    IsActive    TINYINT NOT NULL DEFAULT 1,
                    CreatedAt   DATETIME NOT NULL DEFAULT GETDATE()
                )""")
            conn.commit()
            print("[startup] Created BOA.ZZZ.Objective")
        else:
            print("[startup] Objective exists — skipping")

        # 311xxxx — KeyResult
        if not _table_exists(conn, "KeyResult"):
            conn.execute("""
                CREATE TABLE BOA.ZZZ.KeyResult (
                    KRID          INT IDENTITY(3110001,1) PRIMARY KEY,
                    ObjectiveID   INT NOT NULL,
                    Title         NVARCHAR(300) NOT NULL,
                    TargetValue   DECIMAL(18,2) NOT NULL DEFAULT 0,
                    AchievedValue DECIMAL(18,2) NOT NULL DEFAULT 0,
                    PipelineValue DECIMAL(18,2) NOT NULL DEFAULT 0,
                    Unit          NVARCHAR(50),
                    CalcMethod    NVARCHAR(20)  NOT NULL DEFAULT 'count',
                    IsActive      TINYINT NOT NULL DEFAULT 1,
                    FOREIGN KEY (ObjectiveID) REFERENCES BOA.ZZZ.Objective(ObjectiveID)
                )""")
            conn.commit()
            print("[startup] Created BOA.ZZZ.KeyResult")
        else:
            print("[startup] KeyResult exists — skipping")
        # 320xxxx — Project
        if not _table_exists(conn, "Project"):
            conn.execute("""
                CREATE TABLE BOA.ZZZ.Project (
                    ProjectID   INT IDENTITY(3200001,1) PRIMARY KEY,
                    ProjectName NVARCHAR(300) NOT NULL,
                    Description NVARCHAR(MAX),
                    Status      NVARCHAR(50) NOT NULL DEFAULT 'Planning',
                    Owner       NVARCHAR(100),
                    StartDate   DATE,
                    Deadline    DATE,
                    ObjectiveID INT,
                    IsActive    TINYINT NOT NULL DEFAULT 1,
                    CreatedAt   DATETIME NOT NULL DEFAULT GETDATE(),
                    UpdatedAt   DATETIME,
                    FOREIGN KEY (ObjectiveID) REFERENCES BOA.ZZZ.Objective(ObjectiveID)
                )""")
            conn.commit()
            print("[startup] Created BOA.ZZZ.Project")
        else:
            print("[startup] Project exists — skipping")

        # 410xxxx — WorkItem
        if not _table_exists(conn, "WorkItem"):
            conn.execute("""
                CREATE TABLE BOA.ZZZ.WorkItem (
                    ItemID      INT IDENTITY(4100001,1) PRIMARY KEY,
                    ParentType  NVARCHAR(10) NOT NULL,
                    ParentID    INT NOT NULL,
                    Title       NVARCHAR(300) NOT NULL,
                    Description NVARCHAR(MAX),
                    Status      NVARCHAR(20) NOT NULL DEFAULT 'not_started',
                    Deadline    DATE,
                    SortOrder   INT NOT NULL DEFAULT 0,
                    IsActive    TINYINT NOT NULL DEFAULT 1,
                    CreatedAt   DATETIME NOT NULL DEFAULT GETDATE(),
                    UpdatedAt   DATETIME
                )""")
            conn.commit()
            print("[startup] Created BOA.ZZZ.WorkItem")
        else:
            print("[startup] WorkItem exists — skipping")

        # 411xxxx — WorkSubItem
        if not _table_exists(conn, "WorkSubItem"):
            conn.execute("""
                CREATE TABLE BOA.ZZZ.WorkSubItem (
                    SubItemID    INT IDENTITY(4110001,1) PRIMARY KEY,
                    ParentItemID INT NOT NULL,
                    Title        NVARCHAR(300) NOT NULL,
                    Status       NVARCHAR(20) NOT NULL DEFAULT 'not_started',
                    Deadline     DATE,
                    SortOrder    INT NOT NULL DEFAULT 0,
                    IsActive     TINYINT NOT NULL DEFAULT 1,
                    CreatedAt    DATETIME NOT NULL DEFAULT GETDATE(),
                    FOREIGN KEY (ParentItemID) REFERENCES BOA.ZZZ.WorkItem(ItemID)
                )""")
            conn.commit()
            print("[startup] Created BOA.ZZZ.WorkSubItem")
        else:
            print("[startup] WorkSubItem exists — skipping")

        # 412xxxx — WorkItemPrerequisite
        if not _table_exists(conn, "WorkItemPrerequisite"):
            conn.execute("""
                CREATE TABLE BOA.ZZZ.WorkItemPrerequisite (
                    LinkID         INT IDENTITY(4120001,1) PRIMARY KEY,
                    ItemID         INT NOT NULL,
                    RequiresItemID INT NOT NULL,
                    IsActive       TINYINT NOT NULL DEFAULT 1,
                    FOREIGN KEY (ItemID)         REFERENCES BOA.ZZZ.WorkItem(ItemID),
                    FOREIGN KEY (RequiresItemID) REFERENCES BOA.ZZZ.WorkItem(ItemID)
                )""")
            conn.commit()
            print("[startup] Created BOA.ZZZ.WorkItemPrerequisite")
        else:
            print("[startup] WorkItemPrerequisite exists — skipping")

        # 413xxxx — WorkItemAssignee
        if not _table_exists(conn, "WorkItemAssignee"):
            conn.execute("""
                CREATE TABLE BOA.ZZZ.WorkItemAssignee (
                    AssigneeID    INT IDENTITY(4130001,1) PRIMARY KEY,
                    ItemID        INT NOT NULL,
                    StakeholderID INT NOT NULL,
                    IsActive      TINYINT NOT NULL DEFAULT 1,
                    FOREIGN KEY (ItemID)        REFERENCES BOA.ZZZ.WorkItem(ItemID),
                    FOREIGN KEY (StakeholderID) REFERENCES BOA.ZZZ.Stakeholder(StakeholderID)
                )""")
            conn.commit()
            print("[startup] Created BOA.ZZZ.WorkItemAssignee")
        else:
            print("[startup] WorkItemAssignee exists — skipping")



        # Syndications Tables
        if not _table_exists(conn, "MainDeals"):
            conn.execute("""
                CREATE TABLE BOA.STR.MainDeals (
                    DealId INT IDENTITY(1,1) PRIMARY KEY,
                    ProductCode NVARCHAR(50),
                    CustomerId INT NOT NULL,
                    FOREIGN KEY (CustomerId) REFERENCES BOA.CUS.Customer(Customerid)
                )""")
            # Create STR schema if it doesn't exist
            try:
                conn.execute("CREATE SCHEMA STR")
                conn.commit()
            except Exception:
                pass # schema might exist
                
            conn.execute("""
                CREATE TABLE BOA.STR.Syndication (
                    DealId INT PRIMARY KEY,
                    Amount FLOAT,
                    Pricing FLOAT,
                    FEC INT,
                    Status NVARCHAR(50),
                    ExpectedDate DATE,
                    FOREIGN KEY (DealId) REFERENCES BOA.STR.MainDeals(DealId)
                )""")
            conn.execute("""
                CREATE TABLE BOA.STR.SyndicationBanks (
                    DealDetailId INT IDENTITY(1,1) PRIMARY KEY,
                    DealId INT NOT NULL,
                    BankName NVARCHAR(200),
                    Amount FLOAT,
                    OfferPricing FLOAT,
                    FOREIGN KEY (DealId) REFERENCES BOA.STR.Syndication(DealId)
                )""")
            conn.commit()
            print("[startup] Created MainDeals, Syndications, SyndicationDetail tables")

            # Clean up old DealType parameters and add 3 dummy records
            conn.execute("DELETE FROM BOA.COR.Parameter WHERE ParamType='DealType'")
            conn.commit()
            
            # Seed 3 dummy syndications
            first_cust_id = conn.execute("SELECT TOP 1 Customerid FROM BOA.CUS.Customer").fetchone()
            if first_cust_id:
                cid = first_cust_id["Customerid"]
                # 1
                conn.execute("INSERT INTO BOA.STR.MainDeals (ProductCode, CustomerId) VALUES ('SYNDICATION', ?)", (cid,))
                did1 = conn.execute("SELECT MAX(DealId) AS id FROM BOA.STR.MainDeals").fetchone()["id"]
                conn.execute("INSERT INTO BOA.STR.Syndication (DealId, Amount, Pricing, FEC, Status, ExpectedDate) VALUES (?, 1000000, 3.5, 2, 'Lead', '2026-12-31')", (did1,))
                conn.execute("INSERT INTO BOA.STR.SyndicationBanks (DealId, BankName, Amount, OfferPricing) VALUES (?, 'Bank A', 500000, 3.6)", (did1,))
                
                # 2
                conn.execute("INSERT INTO BOA.STR.MainDeals (ProductCode, CustomerId) VALUES ('SYNDICATION', ?)", (cid,))
                did2 = conn.execute("SELECT MAX(DealId) AS id FROM BOA.STR.MainDeals").fetchone()["id"]
                conn.execute("INSERT INTO BOA.STR.Syndication (DealId, Amount, Pricing, FEC, Status, ExpectedDate) VALUES (?, 5000000, 4.0, 1, 'Proposal', '2026-10-15')", (did2,))
                conn.execute("INSERT INTO BOA.STR.SyndicationBanks (DealId, BankName, Amount, OfferPricing) VALUES (?, 'Bank B', 2000000, 4.1)", (did2,))
                
                # 3
                conn.execute("INSERT INTO BOA.STR.MainDeals (ProductCode, CustomerId) VALUES ('SYNDICATION', ?)", (cid,))
                did3 = conn.execute("SELECT MAX(DealId) AS id FROM BOA.STR.MainDeals").fetchone()["id"]
                conn.execute("INSERT INTO BOA.STR.Syndication (DealId, Amount, Pricing, FEC, Status, ExpectedDate) VALUES (?, 250000, 2.5, 3, 'Won', '2026-08-01')", (did3,))
                conn.execute("INSERT INTO BOA.STR.SyndicationBanks (DealId, BankName, Amount, OfferPricing) VALUES (?, 'Bank C', 250000, 2.5)", (did3,))
                
                conn.commit()
                print("[startup] Seeded 3 dummy Syndications")
        else:
            print("[startup] MainDeals exists — skipping")

    except Exception as e:
        print(f"[startup] Platform migration error (non-fatal): {e}")
        import traceback; traceback.print_exc()
    finally:
        conn.close()





# ── Products ─────────────────────────────────────────────────────────────────────────

@app.route("/products")
def products_list():
    conn = get_db()
    products = conn.execute(
        "SELECT p.*, "
        "  (SELECT COUNT(*) FROM BOA.STR.MainDeals d WHERE d.ProductCode=p.ProductCode) AS deal_count "
        "FROM BOA.COR.Product p WHERE p.IsActive=1 ORDER BY p.ProductName"
    ).fetchall()
    conn.close()
    return render_template("products.html", products=products)


@app.route("/api/products", methods=["POST"])
def api_product_create():
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "INSERT INTO BOA.COR.Product (ProductCode,ProductName) VALUES (?,?)",
        (data.get("code",""), data.get("name",""))
    )
    conn.commit()
    pid = conn.execute("SELECT MAX(ProductID) AS id FROM BOA.COR.Product").fetchone()["id"]
    conn.close()
    return jsonify({"ok": True, "product_id": pid})


@app.route("/api/products/<int:pid>", methods=["PATCH"])
def api_product_update(pid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "UPDATE BOA.COR.Product SET ProductName=? WHERE ProductID=?",
        (data.get("name"), pid)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/products/<int:pid>", methods=["DELETE"])
def api_product_delete(pid):
    conn = get_db()
    conn.execute("UPDATE BOA.COR.Product SET IsActive=0 WHERE ProductID=?", (pid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


# ── Product Detail & Documents ────────────────────────────────────────────────────────

@app.route("/products/<int:pid>")
def product_detail(pid):
    conn = get_db()
    product = conn.execute(
        "SELECT * FROM BOA.COR.Product WHERE ProductID=? AND IsActive=1", (pid,)
    ).fetchone()
    if not product:
        conn.close()
        return redirect(url_for("products_list"))
    documents = conn.execute(
        "SELECT pd.*, cu.Username AS UploaderName "
        "FROM BOA.COR.ProductDocument pd "
        "LEFT JOIN BOA.COR.[User] cu ON pd.UploadedBy = cu.Username "
        "WHERE pd.ProductID=? AND pd.IsActive=1 "
        "ORDER BY pd.UploadedAt DESC",
        (pid,)
    ).fetchall()
    doc_type_map = get_param_map("PRODDOC", conn)
    conn.close()
    return render_template(
        "product_detail.html",
        product=product,
        documents=documents,
        doc_type_map=doc_type_map
    )


@app.route("/products/<int:pid>/documents", methods=["POST"])
def product_doc_upload(pid):
    conn = get_db()
    product = conn.execute(
        "SELECT ProductID FROM BOA.COR.Product WHERE ProductID=? AND IsActive=1", (pid,)
    ).fetchone()
    if not product:
        conn.close()
        return jsonify({"ok": False, "error": "Product not found"}), 404

    doc_name  = request.form.get("doc_name", "").strip()
    doc_type  = request.form.get("doc_type", "").strip()
    file      = request.files.get("file")

    if not doc_name or not doc_type or not file or file.filename == "":
        conn.close()
        return jsonify({"ok": False, "error": "Missing required fields"}), 400

    original_filename = secure_filename(file.filename)
    file_ext = os.path.splitext(original_filename)[1].lower()
    # Build a unique storage filename: pid_timestamp_original
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    stored_filename = f"{pid}_{timestamp}_{original_filename}"
    save_path = os.path.join(PRODUCT_DOCS_FOLDER, stored_filename)
    file.save(save_path)

    uploader = session.get("username", "system")
    conn.execute(
        "INSERT INTO BOA.COR.ProductDocument "
        "(ProductID, DocName, DocTypeCode, FileName, FileExt, UploadedBy) "
        "VALUES (?,?,?,?,?,?)",
        (pid, doc_name, int(doc_type), stored_filename, file_ext.lstrip("."), uploader)
    )
    conn.commit()
    conn.close()
    return redirect(url_for("product_detail", pid=pid))


@app.route("/api/products/<int:pid>/documents/<int:doc_id>", methods=["PATCH"])
def api_product_doc_edit(pid, doc_id):
    data = request.get_json(silent=True) or {}
    doc_name = data.get("doc_name", "").strip()
    doc_type = data.get("doc_type")
    if not doc_name or not doc_type:
        return jsonify({"ok": False, "error": "Missing fields"}), 400
    conn = get_db()
    conn.execute(
        "UPDATE BOA.COR.ProductDocument SET DocName=?, DocTypeCode=? "
        "WHERE DocID=? AND ProductID=? AND IsActive=1",
        (doc_name, int(doc_type), doc_id, pid)
    )
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/products/<int:pid>/documents/<int:doc_id>", methods=["DELETE"])
def api_product_doc_delete(pid, doc_id):
    conn = get_db()
    row = conn.execute(
        "SELECT FileName FROM BOA.COR.ProductDocument WHERE DocID=? AND ProductID=? AND IsActive=1",
        (doc_id, pid)
    ).fetchone()
    if row:
        conn.execute(
            "UPDATE BOA.COR.ProductDocument SET IsActive=0 WHERE DocID=?", (doc_id,)
        )
        conn.commit()
        file_path = os.path.join(PRODUCT_DOCS_FOLDER, row["FileName"])
        if os.path.exists(file_path):
            os.remove(file_path)
    conn.close()
    return jsonify({"ok": True})


@app.route("/products/<int:pid>/documents/<int:doc_id>/open")
def product_doc_open(pid, doc_id):
    conn = get_db()
    row = conn.execute(
        "SELECT FileName, DocName, FileExt FROM BOA.COR.ProductDocument "
        "WHERE DocID=? AND ProductID=? AND IsActive=1",
        (doc_id, pid)
    ).fetchone()
    conn.close()
    if not row:
        return "File not found", 404
    file_path = os.path.join(PRODUCT_DOCS_FOLDER, row["FileName"])
    if not os.path.exists(file_path):
        return "File not found on disk", 404
    download_name = f"{row['DocName']}.{row['FileExt']}"
    return send_file(file_path, as_attachment=False, download_name=download_name)


def _recalc_all_krs(conn):
    """Auto-recalculate AchievedValue for all active auto-measurement KRs."""
    krs = conn.execute(
        "SELECT * FROM BOA.ZZZ.KeyResult WHERE IsActive=1 AND MeasurementType IN ('product','project')"
    ).fetchall()
    for kr in krs:
        new_val = None
        try:
            if kr["MeasurementType"] == "product" and kr["LinkedProductCode"] and kr["LinkedStatusCodes"]:
                import json
                status_codes = json.loads(kr["LinkedStatusCodes"])
                if status_codes:
                    placeholders = ",".join(["?" for _ in status_codes])
                    params = [kr["LinkedProductCode"]] + status_codes
                    row = conn.execute(
                        f"SELECT COALESCE(SUM(s.Amount),0) AS total "
                        f"FROM BOA.STR.MainDeals m "
                        f"JOIN BOA.STR.Syndication s ON m.DealId = s.DealId "
                        f"WHERE m.ProductCode=? AND s.Status IN ({placeholders})",
                        params
                    ).fetchone()
                    new_val = float(row["total"]) if row else 0.0

            elif kr["MeasurementType"] == "project" and kr["LinkedProjectID"]:
                # Project % done = closed work items / total work items * TargetValue
                total_row = conn.execute(
                    "SELECT COUNT(*) AS cnt FROM BOA.ZZZ.WorkItem WHERE ParentType='project' AND ParentID=? AND IsActive=1",
                    (kr["LinkedProjectID"],)
                ).fetchone()
                done_row = conn.execute(
                    "SELECT COUNT(*) AS cnt FROM BOA.ZZZ.WorkItem WHERE ParentType='project' AND ParentID=? AND IsActive=1 AND Status='done'",
                    (kr["LinkedProjectID"],)
                ).fetchone()
                total_cnt = total_row["cnt"] if total_row else 0
                done_cnt  = done_row["cnt"]  if done_row  else 0
                pct = (done_cnt / total_cnt * 100.0) if total_cnt > 0 else 0.0
                new_val = round(pct * float(kr["TargetValue"]) / 100.0, 2)

            if new_val is not None:
                conn.execute(
                    "UPDATE BOA.ZZZ.KeyResult SET AchievedValue=? WHERE KRID=?",
                    (new_val, kr["KRID"])
                )
        except Exception as e:
            print(f"[recalc] KR {kr['KRID']} error: {e}")
    conn.commit()


@app.route("/okrs")
def okrs_list():
    conn = get_db()
    _recalc_all_krs(conn)
    objectives = conn.execute("SELECT * FROM BOA.ZZZ.Objective WHERE IsActive=1 ORDER BY Period DESC, ObjectiveID").fetchall()
    krs        = conn.execute("SELECT * FROM BOA.ZZZ.KeyResult WHERE IsActive=1 ORDER BY ObjectiveID, KRID").fetchall()
    products   = conn.execute("SELECT * FROM BOA.COR.Product WHERE IsActive=1").fetchall()
    projects   = conn.execute("SELECT ProjectID, ProjectName FROM BOA.ZZZ.Project WHERE IsActive=1 ORDER BY ProjectName").fetchall()
    status_map = get_param_map("Status", conn)
    conn.close()
    return render_template("okrs.html", objectives=objectives, krs=krs,
                           products=products, projects=projects, status_map=status_map)


@app.route("/api/okrs/objectives", methods=["POST"])
def api_objective_create():
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute("INSERT INTO BOA.ZZZ.Objective (Title,Description,Period,Owner) VALUES (?,?,?,?)",
                 (data.get("title",""), data.get("description",""), data.get("period",""), session.get("username","")))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/okrs/objectives/<int:oid>", methods=["DELETE"])
def api_objective_delete(oid):
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.Objective SET IsActive=0 WHERE ObjectiveID=?", (oid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/okrs/krs", methods=["POST"])
def api_kr_create():
    import json as _json
    data = request.get_json(silent=True) or {}
    mtype           = data.get("measurement_type", "manual")
    linked_product  = data.get("linked_product_code") or None
    linked_statuses = _json.dumps(data.get("linked_status_codes", [])) if data.get("linked_status_codes") else None
    linked_project  = data.get("linked_project_id") or None
    conn = get_db()
    conn.execute(
        "INSERT INTO BOA.ZZZ.KeyResult "
        "(ObjectiveID,Title,TargetValue,Unit,CalcMethod,MeasurementType,LinkedProductCode,LinkedStatusCodes,LinkedProjectID) "
        "VALUES (?,?,?,?,?,?,?,?,?)",
        (data.get("objective_id"), data.get("title",""), float(data.get("target",0)),
         data.get("unit",""), data.get("calc_method","manual"),
         mtype, linked_product, linked_statuses, linked_project)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/okrs/krs/<int:krid>", methods=["PATCH"])
def api_kr_update(krid):
    import json as _json
    data = request.get_json(silent=True) or {}
    conn = get_db()
    kr = conn.execute("SELECT * FROM BOA.ZZZ.KeyResult WHERE KRID=?", (krid,)).fetchone()
    if not kr:
        conn.close()
        return jsonify({"ok": False, "error": "KR not found"}), 404

    mtype = data.get("measurement_type", kr["MeasurementType"] or "manual")

    # For manual type, also accept direct achieved/pipeline values
    achieved  = float(data.get("achieved",  kr["AchievedValue"]  or 0))
    pipeline  = float(data.get("pipeline",  kr["PipelineValue"]  or 0))

    linked_product  = data.get("linked_product_code") or None
    linked_statuses = _json.dumps(data.get("linked_status_codes", [])) if data.get("linked_status_codes") else None
    linked_project  = data.get("linked_project_id") or None

    conn.execute(
        "UPDATE BOA.ZZZ.KeyResult SET "
        "Title=?, TargetValue=?, Unit=?, CalcMethod=?, "
        "MeasurementType=?, LinkedProductCode=?, LinkedStatusCodes=?, LinkedProjectID=?, "
        "AchievedValue=?, PipelineValue=? "
        "WHERE KRID=?",
        (
            data.get("title", kr["Title"]),
            float(data.get("target", kr["TargetValue"] or 0)),
            data.get("unit", kr["Unit"] or ""),
            data.get("calc_method", kr["CalcMethod"] or "manual"),
            mtype, linked_product, linked_statuses, linked_project,
            achieved, pipeline,
            krid
        )
    )
    conn.commit()
    # If now auto, recalc immediately
    if mtype in ("product", "project"):
        _recalc_all_krs(conn)
    conn.close()
    return jsonify({"ok": True})



@app.route("/api/okrs/krs/<int:krid>", methods=["DELETE"])
def api_kr_delete(krid):
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.KeyResult SET IsActive=0 WHERE KRID=?", (krid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/okrs/recalculate", methods=["POST"])
def api_okrs_recalculate():
    conn = get_db()
    _recalc_all_krs(conn)
    conn.close()
    return jsonify({"ok": True})


# ── Projects ─────────────────────────────────────────────────────────────────────────

@app.route("/projects")
def projects_list():
    conn = get_db()
    projects = conn.execute(
        "SELECT p.*, o.Title AS ObjTitle, "
        "  (SELECT COUNT(*) FROM BOA.ZZZ.WorkItem w WHERE w.ParentType='project' AND w.ParentID=p.ProjectID AND w.IsActive=1) AS total_items,"
        "  (SELECT COUNT(*) FROM BOA.ZZZ.WorkItem w WHERE w.ParentType='project' AND w.ParentID=p.ProjectID AND w.IsActive=1 AND w.Status='done') AS done_items "
        "FROM BOA.ZZZ.Project p LEFT JOIN BOA.ZZZ.Objective o ON p.ObjectiveID=o.ObjectiveID "
        "WHERE p.IsActive=1 ORDER BY p.CreatedAt DESC"
    ).fetchall()
    objectives = conn.execute("SELECT ObjectiveID, Title FROM BOA.ZZZ.Objective WHERE IsActive=1 ORDER BY Title").fetchall()
    conn.close()
    return render_template("projects.html", projects=projects, objectives=objectives)


@app.route("/projects/<int:project_id>")
def project_detail(project_id):
    conn = get_db()
    project = conn.execute(
        "SELECT p.*, o.Title AS ObjTitle FROM BOA.ZZZ.Project p "
        "LEFT JOIN BOA.ZZZ.Objective o ON p.ObjectiveID=o.ObjectiveID "
        "WHERE p.ProjectID=? AND p.IsActive=1", (project_id,)
    ).fetchone()
    if not project:
        conn.close()
        return "Project not found", 404
    items, prereq_map, subitems_map, assignees_map = _load_backlog(conn, "project", project_id)
    stakeholders = conn.execute("SELECT StakeholderID, FullName, Organization FROM BOA.ZZZ.Stakeholder WHERE IsActive=1 ORDER BY FullName").fetchall()
    users = conn.execute("SELECT id, username, surname FROM BOA.COR.[User] ORDER BY username").fetchall()
    conn.close()
    return render_template("project_detail.html", project=project, items=items, prereq_map=prereq_map, subitems_map=subitems_map, assignees_map=assignees_map, stakeholders=stakeholders, users=users)


@app.route("/api/projects", methods=["POST"])
def api_project_create():
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "INSERT INTO BOA.ZZZ.Project (ProjectName,Description,Status,Owner,StartDate,Deadline,ObjectiveID) VALUES (?,?,?,?,?,?,?)",
        (data.get("name",""), data.get("description",""), data.get("status","Planning"),
         session.get("username",""), data.get("start_date") or None, data.get("deadline") or None, data.get("objective_id") or None)
    )
    conn.commit()
    pid = conn.execute("SELECT MAX(ProjectID) AS id FROM BOA.ZZZ.Project").fetchone()["id"]
    conn.close()
    return jsonify({"ok": True, "project_id": pid})


@app.route("/api/projects/<int:pid>", methods=["PATCH"])
def api_project_update(pid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "UPDATE BOA.ZZZ.Project SET ProjectName=?,Description=?,Status=?,Deadline=?,ObjectiveID=?,UpdatedAt=GETDATE() WHERE ProjectID=?",
        (data.get("name"), data.get("description"), data.get("status"), data.get("deadline") or None, data.get("objective_id") or None, pid)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/projects/<int:pid>", methods=["DELETE"])
def api_project_delete(pid):
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.Project SET IsActive=0 WHERE ProjectID=?", (pid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


# ── Work Items (shared) ───────────────────────────────────────────────────────────────

def _load_backlog(conn, parent_type, parent_id):
    items = conn.execute(
        "SELECT * FROM BOA.ZZZ.WorkItem WHERE ParentType=? AND ParentID=? AND IsActive=1 ORDER BY SortOrder, Deadline, ItemID",
        (parent_type, parent_id)
    ).fetchall()
    item_ids = [i["ItemID"] for i in items]
    prereq_map, subitems_map, assignees_map = {}, {}, {}
    if item_ids:
        ph = ",".join("?" * len(item_ids))
        for p in conn.execute(f"SELECT ItemID,RequiresItemID FROM BOA.ZZZ.WorkItemPrerequisite WHERE ItemID IN ({ph}) AND IsActive=1", item_ids).fetchall():
            prereq_map.setdefault(p["ItemID"], []).append(p["RequiresItemID"])
        for s in conn.execute(f"SELECT * FROM BOA.ZZZ.WorkSubItem WHERE ParentItemID IN ({ph}) AND IsActive=1 ORDER BY SortOrder,SubItemID", item_ids).fetchall():
            subitems_map.setdefault(s["ParentItemID"], []).append(dict(s))
        for a in conn.execute(
            f"SELECT wa.ItemID, COALESCE(s.FullName, u.username + ' ' + ISNULL(u.surname, '')) AS AssigneeName "
            f"FROM BOA.ZZZ.WorkItemAssignee wa "
            f"LEFT JOIN BOA.ZZZ.Stakeholder s ON wa.StakeholderID = s.StakeholderID "
            f"LEFT JOIN BOA.COR.[User] u ON wa.UserID = u.id "
            f"WHERE wa.ItemID IN ({ph}) AND wa.IsActive=1", item_ids
        ).fetchall():
            assignees_map.setdefault(a["ItemID"], []).append(a["AssigneeName"])
    return items, prereq_map, subitems_map, assignees_map


@app.route("/api/workitems", methods=["POST"])
def api_workitem_create():
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute(
        "INSERT INTO BOA.ZZZ.WorkItem (ParentType,ParentID,Title,Description,Deadline,SortOrder) VALUES (?,?,?,?,?,?)",
        (data.get("parent_type"), data.get("parent_id"), data.get("title",""), data.get("description",""), data.get("deadline") or None, data.get("sort_order",0))
    )
    conn.commit()
    iid = conn.execute("SELECT MAX(ItemID) AS id FROM BOA.ZZZ.WorkItem").fetchone()["id"]
    assignees = data.get("assignees", [])
    if assignees:
        for a in assignees:
            if a.startswith("U-"):
                conn.execute("INSERT INTO BOA.ZZZ.WorkItemAssignee (ItemID, UserID) VALUES (?, ?)", (iid, int(a[2:])))
            elif a.startswith("S-"):
                conn.execute("INSERT INTO BOA.ZZZ.WorkItemAssignee (ItemID, StakeholderID) VALUES (?, ?)", (iid, int(a[2:])))
        conn.commit()
    conn.close()
    return jsonify({"ok": True, "item_id": iid})


@app.route("/api/workitems/<int:iid>", methods=["PATCH"])
def api_workitem_update(iid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.WorkItem SET Title=?,Description=?,Status=?,Deadline=?,UpdatedAt=GETDATE() WHERE ItemID=?",
                 (data.get("title"), data.get("description"), data.get("status"), data.get("deadline") or None, iid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/workitems/<int:iid>/status", methods=["PATCH"])
def api_workitem_status(iid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.WorkItem SET Status=?,UpdatedAt=GETDATE() WHERE ItemID=?", (data.get("status"), iid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/workitems/<int:iid>", methods=["DELETE"])
def api_workitem_delete(iid):
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.WorkItem SET IsActive=0 WHERE ItemID=?", (iid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/workitems/<int:iid>/prerequisites", methods=["POST"])
def api_workitem_add_prereq(iid):
    data = request.get_json(silent=True) or {}
    req_id = data.get("requires_item_id")
    conn = get_db()
    if not conn.execute("SELECT LinkID FROM BOA.ZZZ.WorkItemPrerequisite WHERE ItemID=? AND RequiresItemID=? AND IsActive=1", (iid, req_id)).fetchone():
        conn.execute("INSERT INTO BOA.ZZZ.WorkItemPrerequisite (ItemID,RequiresItemID) VALUES (?,?)", (iid, req_id))
        conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/workitems/<int:iid>/prerequisites/<int:req_id>", methods=["DELETE"])
def api_workitem_remove_prereq(iid, req_id):
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.WorkItemPrerequisite SET IsActive=0 WHERE ItemID=? AND RequiresItemID=?", (iid, req_id))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/subitems", methods=["POST"])
def api_subitem_create():
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute("INSERT INTO BOA.ZZZ.WorkSubItem (ParentItemID,Title,Deadline,SortOrder) VALUES (?,?,?,?)",
                 (data.get("parent_item_id"), data.get("title",""), data.get("deadline") or None, data.get("sort_order",0)))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/subitems/<int:sid>/status", methods=["PATCH"])
def api_subitem_status(sid):
    data = request.get_json(silent=True) or {}
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.WorkSubItem SET Status=? WHERE SubItemID=?", (data.get("status"), sid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/subitems/<int:sid>", methods=["DELETE"])
def api_subitem_delete(sid):
    conn = get_db()
    conn.execute("UPDATE BOA.ZZZ.WorkSubItem SET IsActive=0 WHERE SubItemID=?", (sid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


# ── Global Backlog ────────────────────────────────────────────────────────────────────────────────────


@app.route("/backlog")
def global_backlog():
    conn = get_db()
    items = conn.execute(
        "SELECT w.*, "
        "  CASE w.ParentType "
        "    WHEN 'project' THEN (SELECT ProjectName FROM BOA.ZZZ.Project WHERE ProjectID=w.ParentID) "
        "    WHEN 'deal' THEN (SELECT TOP 1 c.CustomerName + ' Deal #'+CAST(d.DealId AS NVARCHAR) "
        "                     FROM BOA.STR.MainDeals d JOIN BOA.CUS.Customer c ON d.CustomerId=c.Customerid "
        "                     WHERE d.DealId=w.ParentID) "
        "  END AS ParentName, "
        "  (SELECT STRING_AGG(COALESCE('S-'+CAST(wa.StakeholderID AS NVARCHAR), 'U-'+CAST(wa.UserID AS NVARCHAR)), ',') "
        "   FROM BOA.ZZZ.WorkItemAssignee wa "
        "   WHERE wa.ItemID = w.ItemID AND wa.IsActive=1) AS AssigneeIDs, "
        "  (SELECT STRING_AGG(COALESCE(s.FullName, u.username + ' ' + ISNULL(u.surname, '')), ', ') "
        "   FROM BOA.ZZZ.WorkItemAssignee wa "
        "   LEFT JOIN BOA.ZZZ.Stakeholder s ON wa.StakeholderID = s.StakeholderID "
        "   LEFT JOIN BOA.COR.[User] u ON wa.UserID = u.id "
        "   WHERE wa.ItemID = w.ItemID AND wa.IsActive=1) AS Assignees "
        "FROM BOA.ZZZ.WorkItem w WHERE w.IsActive=1 AND w.Status != 'done' "
        "ORDER BY w.Deadline ASC, w.SortOrder ASC, w.ItemID ASC"
    ).fetchall()
    stakeholders = conn.execute("SELECT StakeholderID, FullName FROM BOA.ZZZ.Stakeholder WHERE IsActive=1 ORDER BY FullName").fetchall()
    conn.close()
    from datetime import datetime as _dt
    return render_template("backlog.html", items=items, stakeholders=stakeholders, now=_dt.utcnow())


try:
    from rag import search_document_hybrid as _rag_search
    _RAG_AVAILABLE = True
except ImportError:
    _RAG_AVAILABLE = False

# Expert structured finance analyst — system prompt
_RAG_SYSTEM_PROMPT = (
    "You are an expert structured finance analyst and Islamic finance compliance "
    "specialist. You must respond strictly in ENGLISH ONLY, regardless of the "
    "language the user types in.\n\n"
    "Read the provided document excerpts carefully. Your task is to synthesize "
    "the rules, conditions, and exceptions found across the excerpts to form a "
    "single, comprehensive compliance opinion.\n\n"
    "Walk through your reasoning step-by-step using this structure:\n"
    "1. GENERAL PRINCIPLE — State the governing rule or principle.\n"
    "2. CONDITIONS FOR PERMISSIBILITY — List the specific requirements that must "
    "be satisfied.\n"
    "3. PROHIBITIONS & DISQUALIFYING FACTORS — Note anything that would render "
    "the transaction impermissible.\n"
    "4. EXCEPTIONS & SPECIAL CIRCUMSTANCES — Identify any exceptions or edge "
    "cases that apply.\n"
    "5. CONCLUSION — Apply the above to the question and give a clear, direct "
    "compliance opinion.\n\n"
    "If you cannot determine the answer from the provided text, explain precisely "
    "which specific rule, parameter, or contract detail is missing from the "
    "context. Do not speculate or guess beyond what the excerpts state."
)


@app.route("/api/ask-document", methods=["POST"])
def ask_document():
    """Hybrid RAG endpoint with SSE streaming generation.

    Pipeline:
      1. BM25 + multi-query semantic search with RRF fusion (synchronous)
      2. qwen3.5:9b generation streamed token-by-token via SSE

    Request body (JSON):
        { "prompt": "user question" }

    Response: text/event-stream
        data: {"token": "..."}                       <- one per token
        data: {"done": true, "queries_used": [...]}  <- final event
        data: {"error": "..."}                       <- on failure
    """
    try:
        import ollama as _ollama
    except ImportError:
        return jsonify({"error": "ollama not installed"}), 500

    data   = request.get_json(silent=True) or {}
    prompt = (data.get("prompt") or "").strip()

    if not prompt:
        return jsonify({"error": "No prompt provided."}), 400

    if not _RAG_AVAILABLE:
        return jsonify({"error": "RAG module unavailable."}), 500

    # ── Phase 1+2: Hybrid retrieval (synchronous before streaming starts) ───────
    try:
        result  = _rag_search(prompt, n_results=20)
        chunks  = result["chunks"]
        queries = result["queries_used"]
    except FileNotFoundError as exc:
        return jsonify({"error": str(exc)}), 503
    except Exception as exc:
        app.logger.error(f"[ask_document] retrieval error: {exc}")
        return jsonify({"error": f"Retrieval failed: {exc}"}), 500

    context      = "\n\n---\n\n".join(chunks) if chunks else "(No relevant context found.)"
    user_message = (
        f"Document excerpts ({len(chunks)} sections retrieved):\n\n"
        f"{context}\n\n---\n\nQuestion: {prompt}"
    )

    # ── Phase 3: Streaming generation ────────────────────────────────────────────
    def generate():
        try:
            stream = _ollama.chat(
                model="qwen3.5:9b",
                messages=[
                    {"role": "system", "content": _RAG_SYSTEM_PROMPT},
                    {"role": "user",   "content": user_message},
                ],
                stream=True,
            )
            for chunk in stream:
                token = chunk["message"]["content"]
                if token:
                    yield f"data: {json.dumps({'token': token})}\n\n"
            yield f"data: {json.dumps({'done': True, 'queries_used': queries})}\n\n"
        except Exception as exc:
            app.logger.error(f"[ask_document] stream error: {exc}")
            yield f"data: {json.dumps({'error': str(exc)})}\n\n"

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


if __name__ == "__main__":
    _ensure_isactive_columns()
    _run_platform_migrations()
    app.run(debug=True, port=5000)
